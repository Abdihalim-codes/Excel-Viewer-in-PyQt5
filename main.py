from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem
import sys
import openpyxl


class Main(QWidget):
    def __init__(self, parent = ..., flags = ...):
        super(Main, self).__init__()
        self.setWindowTitle("Load Excel Data To QTableWidget")

        layout = QVBoxLayout()
        self.setLayout(layout)

        self.table_widget = QTableWidget()
        layout.addWidget(self.table_widget)

        self.load_data()

    def load_data(self):
        path = "C:/Users/cabdo/Desktop/Python-GUI-Projects/Excel Viewer/list_countries_world.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        self.table_widget.setRowCount(sheet.max_row)
        self.table_widget.setColumnCount(sheet.max_column)


        list_values = list(sheet.values)
        self.table_widget.setHorizontalHeaderLabels(list_values[0])

        row_index = 0
        for value_tuple in list_values[1:]:
            print(value_tuple)
            col_index = 0
            for value in value_tuple:
                self.table_widget.setItem(row_index ,col_index , QTableWidgetItem(str(value)))
                col_index += 1
            row_index += 1


            

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = Main()
    window.showMaximized()
    app.exec_()